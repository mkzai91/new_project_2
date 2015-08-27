from django.utils.encoding import smart_unicode, force_unicode, smart_str
from django.shortcuts import render_to_response,render
from django.template import RequestContext
from bid import models, forms
from decimal import Decimal
from django.http import HttpResponse, Http404,HttpResponseRedirect
import datetime
import pyexcel.ext.xls # pip install pyexcel-xls
#from django.db import models

from models import WorkSheet
from .forms import UploadWorkSheetForm, UploadWorkSheetForm2, RegisterForm
from django.conf import settings
from django import http



def set_cookie(response, key, value, days_expire = 7):
  if days_expire is None:
    max_age = 365 * 24 * 60 * 60  #one year
  else:
    max_age = days_expire * 24 * 60 * 60
  expires = datetime.datetime.strftime(datetime.datetime.utcnow() + datetime.timedelta(seconds=max_age), "%a, %d-%b-%Y %H:%M:%S GMT")
  response.set_cookie(key, value, max_age=max_age, expires=expires, domain=settings.SESSION_COOKIE_DOMAIN, secure=settings.SESSION_COOKIE_SECURE or None)


def check_login(request):
    value= None
    if request.COOKIES.has_key( 'key1' ):
        value = request.COOKIES[ 'key1' ]
    return value

def mainview(request):
    template_name= "main.html"
    msg= ""

    ranking = models.Product.objects.all().order_by("id")
    status=check_login(request)
    if not status:
        return HttpResponseRedirect('/bid/login' )
    status=models.Member.objects.get(username__exact=status)
    if (request.GET.get("search")):
        for rankings in ranking:
            if (request.GET.get("search_product","")) not in rankings.name:
                ranking=ranking.exclude(name=rankings.name)
        if (request.GET.get("dropdown","")) == 'name':
            ranking = ranking.order_by("name")
        elif (request.GET.get("dropdown",""))=='bid_price':
            ranking = ranking.order_by("bid_price")
        elif (request.GET.get("dropdown",""))=='popularity':
            ranking = ranking.order_by("-total_view")
        else:
            ranking = ranking.order_by("expire_date")
    
    new_prices = models.Product.objects.all()
    for new_price in new_prices:
        if request.GET.get(str(new_price.id)+"_bid"):
            user_input_add = request.GET.get(str(new_price.id)+'_try_price', '')
            new_price_2 = models.Product.objects.get(pk=new_price.id)#only get one item
            try:
                if float(user_input_add) > float(new_price_2.bid_price):
                    new_price_2.bid_price = Decimal(user_input_add)
                    msg = " Price Updated "
                    new_price_2.save()
            except:
                msg = "please enter you bid price"
                
    buyproducts = models.Product.objects.all()
    for buyproduct in buyproducts:
        if request.GET.get(str(buyproduct.id)+"_pay"):
            print 111
            product = models.Product.objects.get(pk=buyproduct.id)#only get one item
            product.buyer= request.COOKIES[ 'key1' ]
            product.save()
            return HttpResponseRedirect('/bid/checkout' )


    context = {
        'rankings': ranking,
        'msg': msg,
        'status' : status
        }
    querys= models.Product.objects.all().values('name','publish_date','price','expire_date','bid_price','total_view')
    if (request.GET.get("addproduct")):
        return HttpResponseRedirect('/bid/addproduct' )
    elif (request.GET.get("showimage")):
        return HttpResponseRedirect('/bid/showimage' )
    elif (request.GET.get("upload")):
        return HttpResponseRedirect('/bid/upload' )
    elif (request.GET.get("excel")):
        return HttpResponse(create_excel(querys,False),content_type='application/vnd.ms-excel')
    else:
        return render(request,template_name,context)


def addproduct(request):
    template_name= "addproduct.html"
    msg= ""
    status=check_login(request)
    if not status:
        return HttpResponseRedirect('/bid/login' )
    status=models.Member.objects.get(username__exact=status)


    if request.method == 'POST':
            form = UploadWorkSheetForm2(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                querys= models.Product.objects.all().values('name','publish_date','price','expire_date','bid_price','total_view')
                create_excel(querys,True)
                return HttpResponseRedirect('/bid' )
            
    else:

        form = UploadWorkSheetForm2()
    context = RequestContext(request,{
        'msg': msg,
        'form':form,
        'status':status,
        })
    return render_to_response(template_name,context)
  

  #  return render_to_response('addproduct.html',c)

def upload(request):
    messages = []
    template = 'upload.html'
    status=check_login(request)
    if not status:
        return HttpResponseRedirect('/bid/login' )
    status=models.Member.objects.get(username__exact=status)
    if request.method == 'POST':
        form = UploadWorkSheetForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()

            return HttpResponseRedirect('/bid' )
    else:
        form = UploadWorkSheetForm()

    return render(request, 'upload.html', {'form': form,'status':status})

	
def showimage(request):
    template_name='showimage.html'
    ranking = models.WorkSheet.objects.all().order_by("id")
    status=check_login(request)
    if not status:
        return HttpResponseRedirect('/bid/login' )
    status=models.Member.objects.get(username__exact=status)
    context = RequestContext(request,{
        'rankings': ranking,
    })
    if (request.GET.get("back")):
        return HttpResponseRedirect('/bid/' )
    else:
        return render_to_response(template_name,{'status':status})

		

def description(request):
    template_name='description.html'
    msg=''
    status=check_login(request)
    if not status:
        return HttpResponseRedirect('/bid/login' )
    ranking = models.Product.objects.all().order_by("id")
    msg=str(request.path)
    temp=msg
    d=1
    while(True):
        msg=temp[-d:]
        if '/' in msg:
            d-=1
            msg=temp[-d:]
            break
        else:
            d+=1
    msg = int(msg)

    if (request.GET.get("back")):
        return HttpResponseRedirect('/bid/' )
    else:
            for rankings in ranking:
                if msg==rankings.id:
                    rankings.total_view+=1
                    rankings.save()

            context = RequestContext(request,{
                'rankings': ranking,
                'msg':msg,
            })
    return render_to_response(template_name,context)

def create_excel(querys=None,save_internal=True):
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook
    cellname=['A','B','C','D','E','F','G','H','I','J','K']
    wb=Workbook()
    ws=wb.create_sheet()
    dest_filename = 'try_book.xlsx'

    try:
        keys = querys[0].keys()
        for count,key in enumerate (keys):
            ws[cellname[count]+'1']=key
        for count1, query in enumerate(querys):
            for count2, key in enumerate(keys):
                ws[cellname[count2]+str(count1+2)]=query[key]
        
        if(save_internal):
            wb.save(filename=dest_filename)
        else:
            return save_virtual_workbook(wb)
        
    except:
        pass


def login(request):
    template_name='login.html'
    msg=''
    if request.COOKIES.has_key( 'key1' ):
        response = HttpResponseRedirect('/bid/login')
        set_cookie(response ,key='key1',value=None, days_expire=0 )
        return response
    if (request.POST.get("Register")):
        return HttpResponseRedirect('/bid/register' )
    elif(request.POST.get("Login")):
        username= request.POST.get("username")
        password= request.POST.get("password")
        user = None
        try :
            user=models.Member.objects.get(username__exact=username)
        except:
            user = None

        if user and password == user.password:
            msg= 'login success'
            response = HttpResponseRedirect('/bid/')
            if (request.POST.get("check")):
                response.set_cookie(key='key1',value=username,expires=30)
            else:
                response.set_cookie(key='key1',value=username,max_age=None)
            return response
        elif not user:
            msg= "invalid id"
        else:
            msg= 'invalid password'
    context = RequestContext(request,{
                'msg':msg,
            })
    return render_to_response(template_name,context)

def register(request):
    template_name='register.html'
    username= request.POST.get("username")
    password= request.POST.get("password")
    password1= request.POST.get("password1")
    name= request.POST.get("name")
    email= request.POST.get("email")
    form = forms.RegisterForm(request,request.POST)
    cleanform = None
    if (request.POST.get("Register")):
        cleanform = 1
    context = RequestContext(request,{
        'username':username,
        'password':password,
        'name' : name,
        'form' : form,
        'email' : email,
        'cleanform' : cleanform,
    })
    if form.is_valid():
        if password == password1:
            m = models.Member(
                username=context['username'],
                password=context['password'],
                name=context['name'],
                email=context['email'],
            )
            m.save()
            return HttpResponseRedirect('/bid/login' )
    if (request.POST.get("back")):
        return HttpResponseRedirect('/bid/login' )
    return render_to_response(template_name,context)

def checkout(request):
    template_name='checkout.html'
    status=check_login(request)
    if not status:
        return HttpResponseRedirect('/bid/login' )
    status=models.Member.objects.get(username__exact=status)
    name=request.COOKIES[ 'key1' ]
    product=models.Product.objects.filter(buyer__exact=name)
    total_price=0
    print product
    for products in product:
        total_price +=products.price
    member=models.Member.objects.get(username__exact=name)
    context = RequestContext(request,{
        'product':product,
        'member' : member,
        'total_price' : total_price,
        'status':status
    })

    if (request.POST.get("order_now")):
        print 1234
        if (request.POST.get('PayOption',"PayPlatform")):
            print 431
            return HttpResponseRedirect('/bid/payment' )



    return render_to_response(template_name,context)

def payment(request):
    template_name='payment.html'
    status=check_login(request)
    print request.POST
    if not status:
        return HttpResponseRedirect('/bid/login' )
    status=models.Member.objects.get(username__exact=status)
    name=request.COOKIES[ 'key1' ]
    product=models.Product.objects.filter(buyer__exact=name)
    total_price=0
    for products in product:
        if request.POST.get(str(products.id)+"_delete"):
            del_product = models.Product.objects.get(pk=products.id)#only get one item
            del_product.buyer = ""
            del_product.save()
            return HttpResponseRedirect('/bid/checkout' )
    for products in product:
        total_price +=products.price
    member=models.Member.objects.get(username__exact=name)
    context = RequestContext(request,{
        'product':product,
        'member' : member,
        'total_price' : total_price,
        'status':status
    })

    return render_to_response(template_name,context)

